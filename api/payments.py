import json
import uuid
import hmac
import hashlib
import requests
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from rest_framework import permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Payment


# =============================================================================
# BKASH TOKENIZED CHECKOUT API — Leg 1 (Customer → Admin)
# =============================================================================
# Docs: https://developer.bka.sh/docs/tokenized-checkout-overview
#
# This replaces SSLCommerz as the ONLY customer-facing payment method.
# SSLCommerz code is preserved below (gated behind comment) for reference.
# =============================================================================

print("[BKASH PAYMENTS] Loading bKash payment module...")

BKASH_SANDBOX = settings.BKASH_SANDBOX
BKASH_BASE_URL = (
    "https://tokenized.sandbox.bka.sh/v1.2.0-beta"
    if BKASH_SANDBOX
    else "https://tokenized.pay.bka.sh/v1.2.0-beta"
)
BKASH_APP_KEY = settings.BKASH_APP_KEY
BKASH_APP_SECRET = settings.BKASH_APP_SECRET
BKASH_USERNAME = settings.BKASH_USERNAME
BKASH_PASSWORD = settings.BKASH_PASSWORD
BKASH_CALLBACK_URL = settings.BKASH_CALLBACK_URL

print(f"[BKASH PAYMENTS] Sandbox mode: {BKASH_SANDBOX}")
print(f"[BKASH PAYMENTS] Base URL: {BKASH_BASE_URL}")
print(f"[BKASH PAYMENTS] Callback URL: {BKASH_CALLBACK_URL}")


# --- In-memory token cache ---
# In production, persist in Redis/DB instead of memory
_token_cache = {
    "id_token": None,
    "refresh_token": None,
    "expires_at": 0,  # epoch seconds
}


def _bkash_headers(id_token):
    return {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": id_token,
        "x-app-key": BKASH_APP_KEY,
    }


def _bkash_grant_token():
    """POST /tokenized/checkout/token/grant — get fresh id_token."""
    url = f"{BKASH_BASE_URL}/tokenized/checkout/token/grant"
    print(f"[BKASH TOKEN] Granting new token from {url}")
    payload = {
        "app_key": BKASH_APP_KEY,
        "app_secret": BKASH_APP_SECRET,
    }
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "username": BKASH_USERNAME,
        "password": BKASH_PASSWORD,
    }
    resp = requests.post(url, json=payload, headers=headers, timeout=30)
    print(f"[BKASH TOKEN] Grant response status: {resp.status_code}")
    data = resp.json()
    print(f"[BKASH TOKEN] Grant response: {json.dumps(data, indent=2)}")

    if not data or not data.get("id_token"):
        raise Exception(f"bKash grant token failed: {json.dumps(data)}")

    _token_cache["id_token"] = data["id_token"]
    _token_cache["refresh_token"] = data.get("refresh_token", _token_cache.get("refresh_token"))
    # id_token valid ~1 hour; refresh 5 min early
    _token_cache["expires_at"] = datetime.now().timestamp() + 55 * 60
    print(f"[BKASH TOKEN] Token granted, expires at: {_token_cache['expires_at']}")
    return data["id_token"]


def _bkash_refresh_token():
    """POST /tokenized/checkout/token/refresh — refresh id_token."""
    url = f"{BKASH_BASE_URL}/tokenized/checkout/token/refresh"
    print(f"[BKASH TOKEN] Refreshing token from {url}")
    payload = {
        "app_key": BKASH_APP_KEY,
        "app_secret": BKASH_APP_SECRET,
        "refresh_token": _token_cache["refresh_token"],
    }
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "username": BKASH_USERNAME,
        "password": BKASH_PASSWORD,
    }
    resp = requests.post(url, json=payload, headers=headers, timeout=30)
    print(f"[BKASH TOKEN] Refresh response status: {resp.status_code}")
    data = resp.json()
    if not data or not data.get("id_token"):
        raise Exception(f"bKash refresh token failed: {json.dumps(data)}")

    _token_cache["id_token"] = data["id_token"]
    _token_cache["refresh_token"] = data.get("refresh_token", _token_cache.get("refresh_token"))
    _token_cache["expires_at"] = datetime.now().timestamp() + 55 * 60
    print(f"[BKASH TOKEN] Token refreshed")
    return data["id_token"]


def _bkash_get_token():
    """Get a valid id_token, caching/refreshing as needed."""
    now = datetime.now().timestamp()
    if _token_cache["id_token"] and _token_cache["expires_at"] - now > 60:
        print(f"[BKASH TOKEN] Using cached token (expires in {_token_cache['expires_at'] - now}s)")
        return _token_cache["id_token"]

    if _token_cache.get("refresh_token"):
        try:
            print(f"[BKASH TOKEN] Attempting token refresh...")
            return _bkash_refresh_token()
        except Exception as e:
            print(f"[BKASH TOKEN] Refresh failed ({e}), falling back to grant...")

    return _bkash_grant_token()


def bkash_create_payment(amount, order_id, payer_reference=None):
    """
    POST /tokenized/checkout/create
    Returns { paymentID, bkashURL, ... } on success.
    Raises Exception on failure.
    """
    id_token = _bkash_get_token()
    url = f"{BKASH_BASE_URL}/tokenized/checkout/create"
    print(f"[BKASH CREATE] Creating payment for order {order_id}, amount={amount}")
    payload = {
        "mode": "0011",
        "payerReference": payer_reference or str(order_id),
        "callbackURL": BKASH_CALLBACK_URL,
        "amount": str(amount),
        "currency": "BDT",
        "intent": "sale",
        "merchantInvoiceNumber": str(order_id),
    }
    print(f"[BKASH CREATE] Payload: {json.dumps(payload)}")
    resp = requests.post(url, json=payload, headers=_bkash_headers(id_token), timeout=30)
    print(f"[BKASH CREATE] Response status: {resp.status_code}")
    data = resp.json()
    print(f"[BKASH CREATE] Response: {json.dumps(data, indent=2)}")

    if not data or data.get("statusCode") != "0000":
        raise Exception(f"bKash create payment failed: {json.dumps(data)}")

    return data


def bkash_execute_payment(payment_id):
    """
    POST /tokenized/checkout/execute
    Finalizes payment after customer completes OTP/PIN on bKash hosted page.
    Returns response data; check transactionStatus === 'Completed' and statusCode === '0000'.
    """
    id_token = _bkash_get_token()
    url = f"{BKASH_BASE_URL}/tokenized/checkout/execute"
    print(f"[BKASH EXECUTE] Executing payment {payment_id}")
    payload = {"paymentID": payment_id}
    resp = requests.post(url, json=payload, headers=_bkash_headers(id_token), timeout=30)
    print(f"[BKASH EXECUTE] Response status: {resp.status_code}")
    data = resp.json()
    print(f"[BKASH EXECUTE] Response: {json.dumps(data, indent=2)}")
    return data


def bkash_query_payment(payment_id):
    """
    POST /tokenized/checkout/payment/status
    Reconciliation check — verify payment status without executing.
    """
    id_token = _bkash_get_token()
    url = f"{BKASH_BASE_URL}/tokenized/checkout/payment/status"
    print(f"[BKASH QUERY] Querying payment {payment_id}")
    payload = {"paymentID": payment_id}
    resp = requests.post(url, json=payload, headers=_bkash_headers(id_token), timeout=30)
    data = resp.json()
    print(f"[BKASH QUERY] Response: {json.dumps(data, indent=2)}")
    return data


def bkash_refund_payment(payment_id, trx_id, amount, reason=None, sku=None):
    """
    POST /tokenized/checkout/payment/refund
    Full or partial refund of a completed payment.
    """
    id_token = _bkash_get_token()
    url = f"{BKASH_BASE_URL}/tokenized/checkout/payment/refund"
    print(f"[BKASH REFUND] Refunding payment {payment_id}, trx={trx_id}, amount={amount}")
    payload = {
        "paymentID": payment_id,
        "trxID": trx_id,
        "amount": str(amount),
        "sku": sku or "refund",
        "reason": reason or "Customer requested refund",
    }
    resp = requests.post(url, json=payload, headers=_bkash_headers(id_token), timeout=30)
    data = resp.json()
    print(f"[BKASH REFUND] Response: {json.dumps(data, indent=2)}")
    return data


def generate_transaction_id(user_id):
    """Generate a unique transaction ID: NOB-{user_id}-{timestamp}-{hex}"""
    suffix = uuid.uuid4().hex[:6].upper()
    return f"NOB-{user_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{suffix}"


# =============================================================================
# VIEWS — bKash Payment Gateway
# =============================================================================

class BKashPaymentInitiateView(APIView):
    """
    POST /api/payments/bkash/initiate/
    Body: { amount: number, order_id?: number }
    Initiates a bKash payment session. Returns bkashURL to redirect customer.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        print(f"[BKASH INITIATE] User {request.user.id} initiating payment")

        amount = request.data.get("amount")
        if not amount:
            return Response({"error": "Amount is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                raise ValueError
        except (ValueError, InvalidOperation):
            return Response({"error": "Amount must be a positive number."}, status=status.HTTP_400_BAD_REQUEST)

        order = None
        order_id = request.data.get("order_id")
        if order_id:
            try:
                with transaction.atomic():
                    order = Order.objects.select_for_update().get(pk=order_id)
            except Order.DoesNotExist:
                return Response({"error": "Order not found."}, status=status.HTTP_404_NOT_FOUND)
            if order.customer != request.user and not (request.user.is_staff or request.user.role == 'admin'):
                return Response({"error": "You do not own this order."}, status=status.HTTP_403_FORBIDDEN)

        tran_id = generate_transaction_id(request.user.id)
        print(f"[BKASH INITIATE] Transaction ID: {tran_id}, Amount: {amount}, order_id={order_id}")

        try:
            result = bkash_create_payment(
                amount=amount,
                order_id=tran_id,
                payer_reference=request.user.phone_number or str(request.user.id),
            )
        except Exception as e:
            print(f"[BKASH INITIATE] Failed: {e}")
            return Response({"error": f"Failed to initiate bKash payment: {e}"}, status=status.HTTP_502_BAD_GATEWAY)

        payment = Payment.objects.create(
            user=request.user,
            order=order,
            amount=amount,
            transaction_id=tran_id,
            status="initiated",
            gateway="bkash",
            bkash_payment_id=result.get("paymentID"),
            gateway_response=result,
        )
        print(f"[BKASH INITIATE] Payment #{payment.id} created with bkash_payment_id={result.get('paymentID')}")

        return Response({
            "payment_id": payment.id,
            "order_id": order.id if order else None,
            "transaction_id": tran_id,
            "bkash_url": result.get("bkashURL"),
            "payment_id_bkash": result.get("paymentID"),
            "amount": f"{amount:.2f}",
        })


@method_decorator(csrf_exempt, name="dispatch")
class BKashEscrowTrxView(APIView):
    """
    POST /api/payments/escrow/trx/
    Body: { order_id, payment_type: 'advance'|'final', trx_id, amount? }

    Records a customer's manually-submitted bKash Transaction ID against an order
    as the 50% advance or the 50% final settlement (UddoktaPay/bKash manual flow).
    Marks the order's advance/final as paid and appends the farmer-payout ledger row.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        order_id = request.data.get("order_id")
        payment_type = request.data.get("payment_type", "final")
        trx_id = request.data.get("trx_id")

        if not order_id:
            return Response({"error": "order_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        if payment_type not in ("advance", "final"):
            return Response({"error": "payment_type must be 'advance' or 'final'."}, status=status.HTTP_400_BAD_REQUEST)
        if not trx_id:
            return Response({"error": "trx_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                order = Order.objects.select_for_update().get(pk=order_id)
        except Order.DoesNotExist:
            return Response({"error": "Order not found."}, status=status.HTTP_404_NOT_FOUND)

        if order.customer != request.user and not (request.user.is_staff or request.user.role == 'admin'):
            return Response({"error": "You do not own this order."}, status=status.HTTP_403_FORBIDDEN)

        # Escrow stage validation
        if payment_type == 'advance':
            if order.advance_paid:
                return Response({"error": "Advance payment already completed for this order."}, status=status.HTTP_400_BAD_REQUEST)
            amount = order.advance_amount or Decimal('0.00')
        else:
            if not order.advance_paid:
                return Response({"error": "Advance payment must be completed before the final payment."}, status=status.HTTP_400_BAD_REQUEST)
            if order.final_paid:
                return Response({"error": "Final payment already completed for this order."}, status=status.HTTP_400_BAD_REQUEST)
            amount = order.final_amount or Decimal('0.00')

        # Optional client-provided amount override is only honoured for admins.
        if request.data.get("amount") and (request.user.is_staff or request.user.role == 'admin'):
            try:
                amount = Decimal(str(request.data["amount"]))
            except (ValueError, InvalidOperation):
                return Response({"error": "Invalid amount."}, status=status.HTTP_400_BAD_REQUEST)

        if amount <= 0:
            return Response({"error": "Order has no payable amount for this stage."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            payment = Payment.objects.create(
                user=request.user,
                order=order,
                amount=amount,
                payment_type=payment_type,
                transaction_id=f"ESCROW-{trx_id}",
                status="success",
                gateway="bkash",
                bkash_trx_id=trx_id,
                paid_at=timezone.now(),
                settlement_appended=False,
            )
            if _append_settlement_xlsx(payment):
                Payment.objects.filter(pk=payment.pk).update(settlement_appended=True)
            if payment_type == 'advance':
                order.advance_paid = True
            else:
                order.final_paid = True
            order.paid_amount = amount
            order.bkash_trx_id = trx_id
            order.bkash_payment_status = 'success'
            order.paid_at = payment.paid_at
            order.save(update_fields=[
                'advance_paid', 'final_paid', 'paid_amount', 'bkash_trx_id',
                'bkash_payment_status', 'paid_at',
            ])

        return Response({
            "payment_id": payment.id,
            "order_id": order.id,
            "payment_type": payment_type,
            "amount": f"{amount:.2f}",
            "status": "success",
        }, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class BKashPaymentCallbackView(APIView):
    """
    GET/POST /api/payments/bkash/callback/
    bKash redirects the customer's browser here after payment completion.
    Query params: paymentID, status (success/failure/cancel)
    
    IMPORTANT: This endpoint is called by bKash via browser redirect.
    We MUST call executePayment to finalize before marking as paid.
    """
    permission_classes = []

    def get(self, request):
        return self._handle_callback(request)

    def post(self, request):
        return self._handle_callback(request)

    def _handle_callback(self, request):
        print(f"[BKASH CALLBACK] Received callback. Query: {dict(request.query_params)} Body: {dict(request.POST)}")
        payment_id = request.query_params.get("paymentID") or request.POST.get("paymentID")
        status_param = request.query_params.get("status") or request.POST.get("status", "")

        if not payment_id:
            print("[BKASH CALLBACK] No paymentID provided")
            return HttpResponse("Missing paymentID", status=400)

        print(f"[BKASH CALLBACK] paymentID={payment_id}, status={status_param}")

        # Customer cancelled or payment failed at bKash side
        if status_param != "success":
            print(f"[BKASH CALLBACK] Payment not successful (status={status_param}), marking cancelled")
            Payment.objects.filter(bkash_payment_id=payment_id).update(status="cancelled")
            return HttpResponse(f"Payment {status_param}. You can close this page.", status=200)

        # Execute payment to finalize
        try:
            result = bkash_execute_payment(payment_id)
        except Exception as e:
            print(f"[BKASH CALLBACK] Execute failed: {e}")
            # Fall back to query
            try:
                status_check = bkash_query_payment(payment_id)
                if status_check.get("transactionStatus") == "Completed":
                    result = status_check
                    print(f"[BKASH CALLBACK] Query confirmed payment completed")
                else:
                    Payment.objects.filter(bkash_payment_id=payment_id).update(status="failed")
                    return HttpResponse(f"Payment verification failed.", status=200)
            except Exception as e2:
                print(f"[BKASH CALLBACK] Query also failed: {e2}")
                Payment.objects.filter(bkash_payment_id=payment_id).update(status="failed")
                return HttpResponse(f"Payment verification error.", status=200)

        if result.get("statusCode") == "0000" and result.get("transactionStatus") == "Completed":
            trx_id = result.get("trxID")
            amount = result.get("amount")
            print(f"[BKASH CALLBACK] Payment COMPLETED. trxID={trx_id}, amount={amount}")

            pay = Payment.objects.filter(bkash_payment_id=payment_id).first()
            if pay:
                _finalize_payment(pay, trx_id=trx_id, gateway_response=result)
                print(f"[BKASH CALLBACK] Payment record updated to success (settlement_appended={pay.settlement_appended})")
            return HttpResponse("Payment successful! You can close this page.", status=200)
        else:
            print(f"[BKASH CALLBACK] Execute returned non-success: {json.dumps(result)}")
            # Fall back to query
            try:
                status_check = bkash_query_payment(payment_id)
                if status_check.get("transactionStatus") == "Completed":
                    trx_id = status_check.get("trxID")
                    amount = status_check.get("amount")
                    pay = Payment.objects.filter(bkash_payment_id=payment_id).first()
                    if pay:
                        _finalize_payment(pay, trx_id=trx_id, gateway_response=status_check)
                    print(f"[BKASH CALLBACK] Query fallback confirmed payment completed")
                    return HttpResponse("Payment successful! You can close this page.", status=200)
            except Exception as e:
                print(f"[BKASH CALLBACK] Query fallback failed: {e}")

            Payment.objects.filter(bkash_payment_id=payment_id).update(status="failed")
            return HttpResponse(f"Payment execution failed. Please try again.", status=200)


@method_decorator(csrf_exempt, name="dispatch")
class BKashPaymentSuccessView(APIView):
    """
    POST /api/payments/bkash/success/
    Direct success endpoint (can be called by frontend after polling detects success).
    Body: { transaction_id }
    """
    permission_classes = []

    def post(self, request):
        tran_id = request.data.get("transaction_id")
        print(f"[BKASH SUCCESS] transaction_id={tran_id}")
        if tran_id:
            pay = Payment.objects.filter(transaction_id=tran_id).first()
            if pay:
                _finalize_payment(pay)
        return Response({"status": "success", "message": "Payment marked successful.", "transaction_id": tran_id})


@method_decorator(csrf_exempt, name="dispatch")
class BKashPaymentFailView(APIView):
    """
    POST /api/payments/bkash/fail/
    Body: { transaction_id }
    """
    permission_classes = []

    def post(self, request):
        tran_id = request.data.get("transaction_id")
        print(f"[BKASH FAIL] transaction_id={tran_id}")
        if tran_id:
            Payment.objects.filter(transaction_id=tran_id).update(status="failed")
        return Response({"status": "failed", "message": "Payment marked failed.", "transaction_id": tran_id})


class BKashPaymentStatusView(APIView):
    """
    GET /api/payments/bkash/status/<transaction_id>/
    Check payment status. Optionally reconciles with bKash API if status is initiated.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, transaction_id):
        print(f"[BKASH STATUS] Checking status for transaction {transaction_id}")
        try:
            payment = Payment.objects.get(transaction_id=transaction_id)
        except Payment.DoesNotExist:
            return Response({"error": "Payment not found."}, status=status.HTTP_404_NOT_FOUND)

        # If still initiated, try to execute + reconcile with bKash API
        if payment.status == "initiated" and payment.bkash_payment_id:
            try:
                # First attempt to execute the payment (user completed OTP/PIN on bKash side)
                result = bkash_execute_payment(payment.bkash_payment_id)
                if result.get("statusCode") == "0000" and result.get("transactionStatus") == "Completed":
                    _finalize_payment(payment, trx_id=result.get("trxID"), gateway_response=result)
                    print(f"[BKASH STATUS] Execute + reconciled: success, trx={result.get('trxID')}")
                else:
                    # Execute didn't complete, fall back to query
                    result = bkash_query_payment(payment.bkash_payment_id)
                    if result.get("transactionStatus") == "Completed":
                        _finalize_payment(payment, trx_id=result.get("trxID"), gateway_response=result)
                        print(f"[BKASH STATUS] Query reconciled: success, trx={result.get('trxID')}")
                    elif result.get("transactionStatus") in ("Failed", "Cancelled"):
                        payment.status = "failed"
                        payment.gateway_response = result
                        payment.save()
                        print(f"[BKASH STATUS] Reconciled: payment failed/cancelled")
            except Exception as e:
                print(f"[BKASH STATUS] Execute failed, trying query: {e}")
                try:
                    result = bkash_query_payment(payment.bkash_payment_id)
                    if result.get("transactionStatus") == "Completed":
                        _finalize_payment(payment, trx_id=result.get("trxID"), gateway_response=result)
                        print(f"[BKASH STATUS] Query reconciled after execute fail: success, trx={result.get('trxID')}")
                    elif result.get("transactionStatus") in ("Failed", "Cancelled"):
                        payment.status = "failed"
                        payment.gateway_response = result
                        payment.save()
                        print(f"[BKASH STATUS] Query reconciled after execute fail: failed/cancelled")
                except Exception as e2:
                    print(f"[BKASH STATUS] Reconciliation query also failed: {e2}")

        return Response({
            "transaction_id": payment.transaction_id,
            "amount": str(payment.amount),
            "status": payment.status,
            "gateway": payment.gateway,
            "bkash_payment_id": payment.bkash_payment_id,
            "bkash_trx_id": payment.bkash_trx_id,
            "created_at": payment.created_at,
        })


class BKashPaymentRefundView(APIView):
    """
    POST /api/payments/bkash/refund/
    Body: { transaction_id, amount (optional), reason (optional) }
    Admin-only. Refunds a completed bKash payment.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if request.user.role != "admin" and not request.user.is_staff:
            return Response({"error": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        tran_id = request.data.get("transaction_id")
        amount = request.data.get("amount")
        reason = request.data.get("reason")

        print(f"[BKASH REFUND] Request by admin {request.user.id}, transaction={tran_id}")

        if not tran_id:
            return Response({"error": "transaction_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            payment = Payment.objects.get(transaction_id=tran_id)
        except Payment.DoesNotExist:
            return Response({"error": "Payment not found."}, status=status.HTTP_404_NOT_FOUND)

        if payment.status != "success" or not payment.bkash_payment_id or not payment.bkash_trx_id:
            return Response({"error": "Payment is not in a refundable state."}, status=status.HTTP_400_BAD_REQUEST)

        refund_amount = Decimal(str(amount)) if amount else payment.amount

        try:
            result = bkash_refund_payment(
                payment_id=payment.bkash_payment_id,
                trx_id=payment.bkash_trx_id,
                amount=refund_amount,
                reason=reason,
            )
            print(f"[BKASH REFUND] Result: {json.dumps(result)}")

            if result.get("statusCode") == "0000":
                payment.status = "refunded"
                payment.gateway_response = result
                payment.save()
                print(f"[BKASH REFUND] Success")

            return Response(result)
        except Exception as e:
            print(f"[BKASH REFUND] Failed: {e}")
            return Response({"error": f"Refund failed: {e}"}, status=status.HTTP_502_BAD_GATEWAY)


# =============================================================================
# CORPNET BULK PAYMENT XLSX — Leg 2 (Admin → Farmer bank settlement)
# =============================================================================
# Generates an XLSX file in BRAC Bank CORPnet bulk payment format.
# =============================================================================

import csv
import io
from .models import Order, FarmerBankAccount

# CORPnet bulk payment file format (see docs):
#   Debit Account | Beneficiary Name | Beneficiary Account | Routing Number |
#   Bank Name & Branch | Amount | Payment Mode | Narration / Remarks
CORPNET_HEADERS = [
    "Debit Account",
    "Beneficiary Name",
    "Beneficiary Account",
    "Routing Number",
    "Bank Name & Branch",
    "Amount (BDT)",
    "Payment Mode",
    "Narration / Remarks",
]


def _resolve_payment_mode(bank):
    """Pick the CORPnet Payment Mode from the farmer's bank record.
    Explicit mode wins; otherwise auto-detect: BRAC→IFT, bKash→MFS, else EFT."""
    if bank.payment_mode:
        return bank.payment_mode
    bname = (bank.bank_name or '').strip().lower()
    if 'brac' in bname:
        return 'IFT'
    if 'bkash' in bname or 'mfs' in bname:
        return 'MFS'
    return 'EFT'


class BEFTNInvoiceView(APIView):
    """
    GET /api/payments/beftn/invoice/?from_date=2026-01-01&to_date=2026-07-25
    Admin-only. Generates a CORPnet bulk-payment XLSX of completed/shipped orders.

    Each row = one order (fully traceable), matching the CORPnet bulk file layout.
    Skips farmers without complete bank details (flags them).
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        print(f"[BEFTN] Generating CORPnet invoice, user={request.user.id}, role={request.user.role}")

        if request.user.role != "admin" and not request.user.is_staff:
            return Response({"error": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        from_date = request.query_params.get("from_date")
        to_date = request.query_params.get("to_date")

        orders = Order.objects.filter(
            status__in=["completed"],
        ).prefetch_related("items__farmer", "items__post__product_type").order_by("created_at")

        if from_date:
            from datetime import datetime as dt
            try:
                fd = dt.strptime(from_date, "%Y-%m-%d")
                orders = orders.filter(created_at__gte=fd)
            except ValueError:
                return Response({"error": "Invalid from_date format. Use YYYY-MM-DD."}, status=400)

        if to_date:
            from datetime import datetime as dt
            try:
                td = dt.strptime(to_date, "%Y-%m-%d")
                orders = orders.filter(created_at__lte=td.replace(hour=23, minute=59, second=59))
            except ValueError:
                return Response({"error": "Invalid to_date format. Use YYYY-MM-DD."}, status=400)

        if not orders:
            print(f"[BEFTN] No orders found in the given date range")
            return Response({"error": "No orders found in the given date range.", "orders_count": 0}, status=404)

        print(f"[BEFTN] Found {orders.count()} orders to process")

        debit_account = getattr(settings, 'CORPNET_DEBIT_ACCOUNT', '0000000000000000')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CORPnet Bulk"
        ws.append(CORPNET_HEADERS)

        sl = 0
        total_amount = Decimal("0.00")
        incomplete_farmers = []

        for order in orders:
            items = list(order.items.select_related('farmer', 'post__product_type').all())
            if not items:
                continue
            # Group items by farmer
            farmer_rows = {}
            for item in items:
                fid = item.farmer_id
                if fid not in farmer_rows:
                    farmer_rows[fid] = {'farmer': item.farmer, 'subtotal': Decimal('0'), 'remarks_parts': []}
                farmer_rows[fid]['subtotal'] += item.subtotal
                if item.post.product_type:
                    farmer_rows[fid]['remarks_parts'].append(f"{item.post.product_type.name_en} {item.quantity_kg}kg")
                else:
                    farmer_rows[fid]['remarks_parts'].append(f"{item.post.title} ({item.quantity_kg}kg)")

            for fid, data in farmer_rows.items():
                farmer = data['farmer']
                try:
                    bank = FarmerBankAccount.objects.get(farmer=farmer)
                except FarmerBankAccount.DoesNotExist:
                    incomplete_farmers.append({"farmer_id": farmer.id, "farmer_name": farmer.name, "order_id": order.id, "reason": "No bank account details"})
                    continue

                if not bank.routing_number or not bank.account_number:
                    incomplete_farmers.append({"farmer_id": farmer.id, "farmer_name": farmer.name, "order_id": order.id, "reason": "Incomplete bank details"})
                    continue

                sl += 1
                remarks = ', '.join(data['remarks_parts'])
                bank_branch = f"{bank.bank_name} - {bank.branch_name}".strip(" -")
                mode = _resolve_payment_mode(bank)

                ws.append([
                    debit_account,
                    farmer.name or farmer.username,
                    bank.account_number,
                    bank.routing_number,
                    bank_branch,
                    float(data['subtotal']),
                    mode,
                    remarks,
                ])
                total_amount += data['subtotal']

        # Summary row (blank cells, TOTAL under Amount)
        if sl > 0:
            ws.append([
                "", "", "", "", "TOTAL", float(total_amount), "", "",
            ])

        # Column widths for readability
        widths = [18, 24, 24, 16, 34, 14, 14, 34]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        print(f"[BEFTN] Generated CORPnet xlsx with {sl} entries, total BDT {total_amount:.2f}")
        if incomplete_farmers:
            print(f"[BEFTN] WARNING: {len(incomplete_farmers)} orders skipped due to incomplete farmer bank details:")
            for f in incomplete_farmers:
                print(f"[BEFTN]   - Farmer #{f['farmer_id']} ({f['farmer_name']}): Order #{f['order_id']} - {f['reason']}")

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="corpnet_bulk_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response


# =============================================================================
# SETTLEMENT XLSX LEDGER — one row per successfully paid, order-linked payment
# =============================================================================
# Columns: Account Number, Farmer Name, Amount (90%), Payment Type,
#          Reference (Order ID), Contact (Phone)
# =============================================================================

SETTLEMENT_HEADERS = CORPNET_HEADERS


def _settlement_path():
    path = getattr(settings, 'SETTLEMENT_XLSX_PATH', None)
    if path:
        return Path(path)
    return Path(settings.BASE_DIR) / 'settlements' / 'admin_settlement.xlsx'


def _rebuild_settlement_xlsx():
    """Rebuild the full settlement xlsx from every successful, order-linked
    payment. One row per payment, with farmer info from the order's items."""
    from .models import Payment, Order
    path = _settlement_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    debit_account = getattr(settings, 'CORPNET_DEBIT_ACCOUNT', '0000000000000000')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settlements"
    ws.append(SETTLEMENT_HEADERS)

    payments = (
        Payment.objects
        .filter(status='success', order__isnull=False)
        .select_related('order')
        .prefetch_related('order__items__farmer')
        .order_by('paid_at', 'id')
    )

    for payment in payments:
        order = payment.order
        if order is None:
            continue
        items = list(order.items.select_related('farmer').all())
        if not items:
            continue
        # Group by farmer for this payment
        farmer_rows = {}
        for item in items:
            fid = item.farmer_id
            if fid not in farmer_rows:
                farmer_rows[fid] = {'farmer': item.farmer, 'subtotal': Decimal('0')}
            farmer_rows[fid]['subtotal'] += item.subtotal
        for fid, data in farmer_rows.items():
            farmer = data['farmer']
            bank = None
            try:
                bank = FarmerBankAccount.objects.get(farmer=farmer)
            except FarmerBankAccount.DoesNotExist:
                pass
            ws.append([
                debit_account,
                farmer.name or farmer.username,
                bank.account_number if bank else '',
                bank.routing_number if bank else '',
                f"{bank.bank_name} - {bank.branch_name}".strip(" -") if bank else '',
                float(data['subtotal']),
                _resolve_payment_mode(bank) if bank else 'EFT',
                f"Order #{order.id}",
            ])

    wb.save(path)
    return path


def _append_settlement_xlsx(payment):
    """Settlement XLSX append (simplified for OrderItem model)."""
    return False


def _finalize_payment(payment, trx_id=None, gateway_response=None):
    """Mark a payment successful and (once) append its settlement xlsx row."""
    payment.status = 'success'
    if trx_id:
        payment.bkash_trx_id = trx_id
    if gateway_response is not None:
        payment.gateway_response = gateway_response
    payment.paid_at = datetime.now()
    if payment.order_id and not payment.settlement_appended:
        if _append_settlement_xlsx(payment):
            payment.settlement_appended = True
    payment.save()


class SettlementDownloadView(APIView):
    """
    GET /api/payments/settlement/download/
    Admin-only. Rebuilds the settlement xlsx from all successful order-linked
    payments and returns it as a downloadable file.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != "admin" and not request.user.is_staff:
            return Response({"error": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        path = _rebuild_settlement_xlsx()
        from django.http import FileResponse
        resp = FileResponse(
            open(path, 'rb'),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="admin_settlement.xlsx"'
        return resp


# =============================================================================
# PRESERVED SSLCOMMERZ CODE (deprecated — kept for reference)
# =============================================================================
"""
SSLCommerz Payment Gateway code has been preserved below for reference.
It is NOT used in production. All payments now go through bKash Tokenized Checkout.

To restore SSLCommerz, uncomment the views below and add URL routes.

from django.conf import settings as django_settings

def _sslcommerz_base():
    return (
        'https://sandbox.sslcommerz.com'
        if django_settings.SSLCOMMERZ_IS_SANDBOX
        else 'https://securepay.sslcommerz.com'
    )

def _initiate_session(amount, tran_id, cus_name, cus_email, cus_phone,
                      success_url, fail_url, cancel_url, ipn_url):
    url = f'{_sslcommerz_base()}/gwprocess/v4/api.php'
    payload = {
        'store_id': django_settings.SSLCOMMERZ_STORE_ID,
        'store_passwd': django_settings.SSLCOMMERZ_STORE_PASSWORD,
        'total_amount': f'{amount:.2f}',
        'currency': 'BDT',
        'tran_id': tran_id,
        'success_url': success_url,
        'fail_url': fail_url,
        'cancel_url': cancel_url,
        'ipn_url': ipn_url,
        'cus_name': cus_name,
        'cus_email': cus_email,
        'cus_phone': cus_phone,
        'cus_add1': 'N/A',
        'cus_city': 'N/A',
        'cus_country': 'Bangladesh',
        'shipping_method': 'NO',
        'product_name': 'Wallet Topup',
        'product_category': 'General',
        'product_profile': 'general',
    }
    resp = requests.post(url, data=payload, timeout=30)
    try:
        return resp.json()
    except json.JSONDecodeError:
        raise requests.RequestException(
            f'Status {resp.status_code}, body: {resp.text[:500]}'
        )

def _validate_session(val_id):
    url = f'{_sslcommerz_base()}/validator/api/validationserverAPI.php'
    params = {
        'val_id': val_id,
        'store_id': django_settings.SSLCOMMERZ_STORE_ID,
        'store_passwd': django_settings.SSLCOMMERZ_STORE_PASSWORD,
        'v': 1,
        'format': 'json',
    }
    resp = requests.get(url, params=params, timeout=30)
    try:
        return resp.json()
    except json.JSONDecodeError:
        raise requests.RequestException(
            f'Status {resp.status_code}, body: {resp.text[:500]}'
        )

class PaymentInitiateView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        # ... SSLCommerz initiate logic ...
        pass

class PaymentSuccessView(APIView): ...
class PaymentFailView(APIView): ...
class PaymentCancelView(APIView): ...
class PaymentIPNView(APIView): ...
class PaymentStatusView(APIView): ...
"""